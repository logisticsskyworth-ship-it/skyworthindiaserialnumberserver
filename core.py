"""
Core logic for the Serial Number Consolidator.
No GUI code here - this module handles the database and Excel import/export
so it can be tested independently of tkinter.
"""
import sqlite3
import os
import hashlib
import secrets
from datetime import datetime
import openpyxl

DB_FILENAME = "serial_numbers.db"

DEFAULT_ADMIN_USERNAME = "Naveen"
DEFAULT_ADMIN_PASSWORD = "Naveen@841"


def get_db_path():
    """DB path is configurable via the DATA_DIR env var (set this to a mounted
    persistent disk when deploying, e.g. Render's Disk feature) so the SQLite
    file survives restarts/redeploys. Falls back to sitting next to this file
    for local/dev use."""
    data_dir = os.environ.get("DATA_DIR")
    if data_dir:
        os.makedirs(data_dir, exist_ok=True)
        return os.path.join(data_dir, DB_FILENAME)
    base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, DB_FILENAME)


def _table_columns(conn, table):
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info({table})")
    return {row[1] for row in cur.fetchall()}


def init_db(db_path=None):
    db_path = db_path or get_db_path()
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS serial_numbers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            serial_number TEXT NOT NULL,
            serial_number_norm TEXT NOT NULL,
            model TEXT,
            invoice_number TEXT,
            vehicle_number TEXT,
            customer_name TEXT,
            entry_date TEXT,
            sheet_name TEXT,
            date_added TEXT,
            source TEXT,
            added_by TEXT,
            status TEXT DEFAULT 'active'
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS duplicates_review (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            serial_number TEXT NOT NULL,
            model TEXT,
            invoice_number TEXT,
            vehicle_number TEXT,
            customer_name TEXT,
            entry_date TEXT,
            sheet_name TEXT,
            date_added TEXT,
            source TEXT,
            added_by TEXT,
            existing_id INTEGER,
            resolved INTEGER DEFAULT 0
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS invoices (
            invoice_number TEXT PRIMARY KEY,
            quantity INTEGER,
            model TEXT,
            vehicle_number TEXT,
            customer_name TEXT,
            entry_date TEXT,
            storage_location TEXT,
            updated_at TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            salt TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user',
            created_at TEXT,
            is_active INTEGER DEFAULT 1
        )
    """)
    cur.execute("CREATE TABLE IF NOT EXISTS models (name TEXT PRIMARY KEY)")
    cur.execute("CREATE TABLE IF NOT EXISTS warehouses (name TEXT PRIMARY KEY)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_serial_norm ON serial_numbers(serial_number_norm)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_invoice_number ON serial_numbers(invoice_number)")
    conn.commit()

    # --- migration for databases created by an earlier version of this app ---
    existing_cols = _table_columns(conn, "serial_numbers")
    for col in ("vehicle_number", "customer_name", "entry_date", "storage_location", "warehouse"):
        if col not in existing_cols:
            cur.execute(f"ALTER TABLE serial_numbers ADD COLUMN {col} TEXT")
    dup_cols = _table_columns(conn, "duplicates_review")
    for col in ("vehicle_number", "customer_name", "entry_date", "storage_location", "warehouse"):
        if col not in dup_cols:
            cur.execute(f"ALTER TABLE duplicates_review ADD COLUMN {col} TEXT")
    inv_cols = _table_columns(conn, "invoices")
    if "storage_location" not in inv_cols:
        cur.execute("ALTER TABLE invoices ADD COLUMN storage_location TEXT")
    if "warehouse" not in inv_cols:
        cur.execute("ALTER TABLE invoices ADD COLUMN warehouse TEXT")
    conn.commit()

    # --- seed the default admin account the very first time the DB is created ---
    cur.execute("SELECT COUNT(*) FROM users WHERE role = 'admin'")
    if cur.fetchone()[0] == 0:
        pwd_hash, salt = _hash_password(DEFAULT_ADMIN_PASSWORD)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            cur.execute("""
                INSERT INTO users (username, password_hash, salt, role, created_at, is_active)
                VALUES (?, ?, ?, 'admin', ?, 1)
            """, (DEFAULT_ADMIN_USERNAME, pwd_hash, salt, now))
            conn.commit()
        except sqlite3.IntegrityError:
            pass
    conn.close()


def _normalize(serial):
    return (serial or "").strip().upper()


# --- User accounts / authentication ---

def _hash_password(password, salt=None):
    if salt is None:
        salt = secrets.token_hex(16)
    pwd_hash = hashlib.pbkdf2_hmac("sha256", (password or "").encode("utf-8"),
                                    salt.encode("utf-8"), 100_000).hex()
    return pwd_hash, salt


def create_user(conn, username, password, role="user"):
    username = (username or "").strip()
    if not username:
        raise ValueError("Username is required.")
    if not password:
        raise ValueError("Password is required.")
    pwd_hash, salt = _hash_password(password)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO users (username, password_hash, salt, role, created_at, is_active)
            VALUES (?, ?, ?, ?, ?, 1)
        """, (username, pwd_hash, salt, role, now))
        conn.commit()
    except sqlite3.IntegrityError:
        raise ValueError(f"Username '{username}' already exists.")
    return cur.lastrowid


def verify_user(conn, username, password):
    """Returns {'id':.., 'username':.., 'role':..} on success, else None."""
    username = (username or "").strip()
    cur = conn.cursor()
    cur.execute("SELECT id, username, password_hash, salt, role, is_active FROM users WHERE username = ?",
                (username,))
    row = cur.fetchone()
    if not row:
        return None
    uid, uname, pwd_hash, salt, role, is_active = row
    if not is_active:
        return None
    check_hash, _ = _hash_password(password, salt)
    if check_hash != pwd_hash:
        return None
    return {"id": uid, "username": uname, "role": role}


def verify_any_admin_password(conn, password):
    """Used to authorize a delete requested by a non-admin user: checks the
    given password against every active admin account."""
    cur = conn.cursor()
    cur.execute("SELECT password_hash, salt FROM users WHERE role = 'admin' AND is_active = 1")
    for pwd_hash, salt in cur.fetchall():
        check_hash, _ = _hash_password(password, salt)
        if check_hash == pwd_hash:
            return True
    return False


def reset_password(conn, username, new_password):
    if not new_password:
        raise ValueError("New password is required.")
    pwd_hash, salt = _hash_password(new_password)
    cur = conn.cursor()
    cur.execute("UPDATE users SET password_hash = ?, salt = ? WHERE username = ?",
                (pwd_hash, salt, (username or "").strip()))
    conn.commit()
    return cur.rowcount > 0


def set_user_active(conn, username, is_active):
    cur = conn.cursor()
    cur.execute("UPDATE users SET is_active = ? WHERE username = ?",
                (1 if is_active else 0, (username or "").strip()))
    conn.commit()
    return cur.rowcount > 0


def list_users(conn):
    cur = conn.cursor()
    cur.execute("SELECT id, username, role, created_at, is_active FROM users ORDER BY username")
    return cur.fetchall()


# --- Model / Warehouse lookup lists (for dropdowns) ---

def list_models(conn):
    cur = conn.cursor()
    cur.execute("SELECT name FROM models ORDER BY name")
    return [r[0] for r in cur.fetchall()]


def add_model(conn, name):
    name = (name or "").strip()
    if not name:
        return
    cur = conn.cursor()
    cur.execute("INSERT OR IGNORE INTO models (name) VALUES (?)", (name,))
    conn.commit()


def list_warehouses(conn):
    cur = conn.cursor()
    cur.execute("SELECT name FROM warehouses ORDER BY name")
    return [r[0] for r in cur.fetchall()]


def add_warehouse(conn, name):
    name = (name or "").strip()
    if not name:
        return
    cur = conn.cursor()
    cur.execute("INSERT OR IGNORE INTO warehouses (name) VALUES (?)", (name,))
    conn.commit()


def import_lookup_lists(conn, filepath):
    """
    Reads an Excel file for Model / Warehouse lists. Accepts either:
      - a single sheet with 'Model' and 'Warehouse' columns (any order), or
      - separate sheets named 'Models' and 'Warehouses' each with one column of names.
    Blank cells are ignored. Returns counts added.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    added_models, added_warehouses = 0, 0

    def norm_header(v):
        return (str(v).strip().lower() if v is not None else "")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = list(next(rows_iter))
        except StopIteration:
            continue
        headers = [norm_header(h) for h in header_row]

        model_idx = next((i for i, h in enumerate(headers) if "model" in h), None)
        warehouse_idx = next((i for i, h in enumerate(headers) if "warehouse" in h), None)

        sheet_lower = sheet_name.strip().lower()
        if model_idx is None and warehouse_idx is None:
            # whole sheet is a single flat list - use the sheet name to decide which list
            if "warehouse" in sheet_lower:
                warehouse_idx = 0
            else:
                model_idx = 0
            # first row wasn't necessarily a header in this case - treat it as data too
            if header_row and header_row[0] not in (None, ""):
                val = str(header_row[0]).strip()
                if warehouse_idx == 0:
                    add_warehouse(conn, val)
                    added_warehouses += 1
                else:
                    add_model(conn, val)
                    added_models += 1

        for row in rows_iter:
            if row is None:
                continue
            if model_idx is not None and model_idx < len(row) and row[model_idx]:
                add_model(conn, str(row[model_idx]).strip())
                added_models += 1
            if warehouse_idx is not None and warehouse_idx < len(row) and row[warehouse_idx]:
                add_warehouse(conn, str(row[warehouse_idx]).strip())
                added_warehouses += 1

    wb.close()
    return {"models_added": added_models, "warehouses_added": added_warehouses}


def find_existing(conn, serial_number):
    norm = _normalize(serial_number)
    cur = conn.cursor()
    cur.execute("SELECT id, serial_number, model, invoice_number, sheet_name, date_added "
                "FROM serial_numbers WHERE serial_number_norm = ? AND status = 'active'", (norm,))
    return cur.fetchone()


# --- Invoice quantity tracking ---

def get_invoice(conn, invoice_number):
    if not invoice_number:
        return None
    cur = conn.cursor()
    cur.execute("SELECT invoice_number, quantity, model, vehicle_number, customer_name, entry_date, "
                "storage_location, warehouse FROM invoices WHERE invoice_number = ?", (invoice_number.strip(),))
    row = cur.fetchone()
    if not row:
        return None
    return {
        "invoice_number": row[0], "quantity": row[1], "model": row[2],
        "vehicle_number": row[3], "customer_name": row[4], "entry_date": row[5],
        "storage_location": row[6], "warehouse": row[7],
    }


def upsert_invoice(conn, invoice_number, quantity=None, model="", vehicle_number="",
                    customer_name="", entry_date="", storage_location="", warehouse=""):
    invoice_number = (invoice_number or "").strip()
    if not invoice_number:
        raise ValueError("Invoice number is required")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO invoices (invoice_number, quantity, model, vehicle_number, customer_name, entry_date,
                               storage_location, warehouse, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(invoice_number) DO UPDATE SET
            quantity=excluded.quantity, model=excluded.model, vehicle_number=excluded.vehicle_number,
            customer_name=excluded.customer_name, entry_date=excluded.entry_date,
            storage_location=excluded.storage_location, warehouse=excluded.warehouse,
            updated_at=excluded.updated_at
    """, (invoice_number, quantity, model, vehicle_number, customer_name, entry_date, storage_location,
          warehouse, now))
    conn.commit()


def count_for_invoice(conn, invoice_number):
    if not invoice_number:
        return 0
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM serial_numbers WHERE invoice_number = ? AND status = 'active'",
                (invoice_number.strip(),))
    return cur.fetchone()[0]


# --- Adding records ---

def add_record(conn, serial_number, model="", invoice_number="", sheet_name="", source="manual",
                added_by="", vehicle_number="", customer_name="", entry_date="", storage_location="",
                warehouse="", enforce_invoice_quantity=True):
    """
    Adds a record.
    Returns one of:
      ('added', new_id)
      ('flagged_duplicate', {'dup_id':.., 'existing_id':.., 'existing_model':.., 'existing_invoice':..,
                              'existing_sheet':.., 'existing_date_added':..})
      ('skipped_empty', None)
      ('exceeds_quantity', {'invoice_number':.., 'quantity':.., 'used':..})
    """
    serial_number = (serial_number or "").strip()
    if not serial_number:
        return ("skipped_empty", None)

    invoice_number = (invoice_number or "").strip()

    if enforce_invoice_quantity and invoice_number:
        inv = get_invoice(conn, invoice_number)
        if inv and inv["quantity"] is not None and inv["quantity"] > 0:
            used = count_for_invoice(conn, invoice_number)
            if used >= inv["quantity"]:
                return ("exceeds_quantity", {
                    "invoice_number": invoice_number, "quantity": inv["quantity"], "used": used
                })

    existing = find_existing(conn, serial_number)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cur = conn.cursor()

    if existing:
        cur.execute("""
            INSERT INTO duplicates_review
            (serial_number, model, invoice_number, vehicle_number, customer_name, entry_date,
             storage_location, warehouse, sheet_name, date_added, source, added_by, existing_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (serial_number, model, invoice_number, vehicle_number, customer_name, entry_date,
              storage_location, warehouse, sheet_name, now, source, added_by, existing[0]))
        conn.commit()
        return ("flagged_duplicate", {
            "dup_id": cur.lastrowid, "existing_id": existing[0], "existing_model": existing[2],
            "existing_invoice": existing[3], "existing_sheet": existing[4], "existing_date_added": existing[5],
        })

    cur.execute("""
        INSERT INTO serial_numbers
        (serial_number, serial_number_norm, model, invoice_number, vehicle_number, customer_name,
         entry_date, storage_location, warehouse, sheet_name, date_added, source, added_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (serial_number, _normalize(serial_number), model, invoice_number, vehicle_number, customer_name,
          entry_date, storage_location, warehouse, sheet_name, now, source, added_by))
    conn.commit()
    return ("added", cur.lastrowid)


# --- Excel import ---

HEADER_ALIASES = {
    "serial_number": ["serial number", "serial no", "serial", "sl no serial", "imei", "imei/serial"],
    "model": ["model", "model no", "model number", "material"],
    "invoice_number": ["invoice number", "invoice no", "invoice", "inv no"],
    "vehicle_number": ["vehicle number", "vehicle no", "truck number", "truck no"],
    "customer_name": ["customer name", "customer", "consignee", "consignee name"],
    "storage_location": ["storage location", "storage loc", "location", "rack location", "rack"],
    "warehouse": ["warehouse", "warehouse name", "wh"],
}


def detect_columns(header_row):
    mapping = {}
    normalized = [(str(v).strip().lower() if v is not None else "") for v in header_row]
    for field, aliases in HEADER_ALIASES.items():
        for idx, header in enumerate(normalized):
            if header in aliases:
                mapping[field] = idx
                break
        if field not in mapping:
            for idx, header in enumerate(normalized):
                if any(alias in header for alias in aliases) and header:
                    mapping[field] = idx
                    break
    return mapping


def import_excel_file(conn, filepath, source_label=None, added_by="", enforce_invoice_quantity=True):
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    summary = {
        "added": 0,
        "flagged_duplicate": 0,
        "skipped_empty": 0,
        "exceeds_quantity": 0,
        "sheets_processed": [],
        "sheets_skipped": [],
    }
    file_label = source_label or os.path.basename(filepath)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            summary["sheets_skipped"].append((sheet_name, "empty sheet"))
            continue

        col_map = detect_columns(list(header_row))
        if "serial_number" not in col_map:
            summary["sheets_skipped"].append((sheet_name, "no serial number column found"))
            continue

        s_idx = col_map["serial_number"]
        m_idx = col_map.get("model")
        i_idx = col_map.get("invoice_number")
        v_idx = col_map.get("vehicle_number")
        c_idx = col_map.get("customer_name")
        l_idx = col_map.get("storage_location")
        w_idx = col_map.get("warehouse")

        def cell(row, idx):
            if idx is None or idx >= len(row) or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        for row in rows_iter:
            if row is None or s_idx >= len(row):
                continue
            serial_val = row[s_idx]
            if serial_val is None or str(serial_val).strip() == "":
                summary["skipped_empty"] += 1
                continue

            status, _info = add_record(
                conn,
                serial_number=str(serial_val).strip(),
                model=cell(row, m_idx),
                invoice_number=cell(row, i_idx),
                vehicle_number=cell(row, v_idx),
                customer_name=cell(row, c_idx),
                storage_location=cell(row, l_idx),
                warehouse=cell(row, w_idx),
                sheet_name=f"{file_label} / {sheet_name}",
                source="upload",
                added_by=added_by,
                enforce_invoice_quantity=enforce_invoice_quantity,
            )
            if status in summary:
                summary[status] += 1

        summary["sheets_processed"].append(sheet_name)

    wb.close()
    return summary


# --- Duplicate review actions ---

def get_pending_duplicates(conn):
    cur = conn.cursor()
    cur.execute("""
        SELECT d.id, d.serial_number, d.model, d.invoice_number, d.sheet_name, d.date_added,
               d.source, d.added_by, d.existing_id,
               e.model, e.invoice_number, e.sheet_name, e.date_added
        FROM duplicates_review d
        LEFT JOIN serial_numbers e ON d.existing_id = e.id
        WHERE d.resolved = 0
        ORDER BY d.id
    """)
    return cur.fetchall()


def resolve_duplicate(conn, dup_id, action):
    cur = conn.cursor()
    cur.execute("""SELECT serial_number, model, invoice_number, vehicle_number, customer_name, entry_date,
                    storage_location, warehouse, sheet_name, date_added, source, added_by, existing_id
                    FROM duplicates_review WHERE id = ?""", (dup_id,))
    row = cur.fetchone()
    if not row:
        return False
    (serial_number, model, invoice_number, vehicle_number, customer_name, entry_date,
     storage_location, warehouse, sheet_name, date_added, source, added_by, existing_id) = row

    if action == "replace" and existing_id:
        cur.execute("""
            UPDATE serial_numbers
            SET serial_number = ?, model = ?, invoice_number = ?, vehicle_number = ?, customer_name = ?,
                entry_date = ?, storage_location = ?, warehouse = ?, sheet_name = ?, date_added = ?
            WHERE id = ?
        """, (serial_number, model, invoice_number, vehicle_number, customer_name, entry_date,
              storage_location, warehouse, sheet_name, date_added, existing_id))
    elif action == "keep_both":
        cur.execute("""
            INSERT INTO serial_numbers
            (serial_number, serial_number_norm, model, invoice_number, vehicle_number, customer_name,
             entry_date, storage_location, warehouse, sheet_name, date_added, source, added_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (serial_number, _normalize(serial_number), model, invoice_number, vehicle_number, customer_name,
              entry_date, storage_location, warehouse, sheet_name, date_added, source, added_by))

    cur.execute("UPDATE duplicates_review SET resolved = 1 WHERE id = ?", (dup_id,))
    conn.commit()
    return True


# --- Search / export ---

def search_master(conn, query="", date_from="", date_to="", limit=500):
    """
    date_from/date_to: 'YYYY-MM-DD' strings, filtered against entry_date when present,
    falling back to the date portion of date_added otherwise. Empty = no bound.
    """
    cur = conn.cursor()
    where = ["status = 'active'"]
    params = []

    if query:
        like = f"%{query.strip()}%"
        where.append("(serial_number LIKE ? OR model LIKE ? OR invoice_number LIKE ? OR "
                      "vehicle_number LIKE ? OR customer_name LIKE ? OR storage_location LIKE ? OR sheet_name LIKE ?)")
        params += [like, like, like, like, like, like, like]

    effective_date = "COALESCE(NULLIF(entry_date, ''), substr(date_added, 1, 10))"
    if date_from:
        where.append(f"{effective_date} >= ?")
        params.append(date_from.strip())
    if date_to:
        where.append(f"{effective_date} <= ?")
        params.append(date_to.strip())

    sql = f"""
        SELECT id, serial_number, model, invoice_number, vehicle_number, customer_name,
               storage_location, warehouse, {effective_date} as eff_date, sheet_name, date_added
        FROM serial_numbers
        WHERE {' AND '.join(where)}
        ORDER BY id DESC LIMIT ?
    """
    params.append(limit)
    cur.execute(sql, params)
    return cur.fetchall()


def count_active(conn):
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM serial_numbers WHERE status = 'active'")
    return cur.fetchone()[0]


def count_pending_duplicates(conn):
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM duplicates_review WHERE resolved = 0")
    return cur.fetchone()[0]


def delete_record(conn, record_id):
    cur = conn.cursor()
    cur.execute("UPDATE serial_numbers SET status = 'deleted' WHERE id = ?", (record_id,))
    conn.commit()


def export_master(conn, out_path, query="", date_from="", date_to=""):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Master_Consolidated"

    headers = ["Sl No", "Serial Number", "Model", "Invoice Number", "Vehicle Number",
               "Customer Name", "Storage Location", "Warehouse", "Date", "Sheet Name", "Scan Time"]
    ws.append(headers)
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for col_idx, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    rows = search_master(conn, query=query, date_from=date_from, date_to=date_to, limit=1000000)
    rows = list(reversed(rows))  # oldest first in export
    for i, r in enumerate(rows, start=1):
        _id, serial, model, invoice, vehicle, customer, storage_location, warehouse, eff_date, sheet, date_added = r
        ws.append([i, serial, model, invoice, vehicle, customer, storage_location, warehouse, eff_date, sheet,
                   date_added])
        for col in range(1, len(headers) + 1):
            ws.cell(row=i + 1, column=col).font = Font(name="Arial")

    widths = [8, 22, 16, 18, 16, 20, 18, 16, 14, 28, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.freeze_panes = "A2"
    wb.save(out_path)
    return len(rows)
